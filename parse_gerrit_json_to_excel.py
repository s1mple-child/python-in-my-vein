'''
Parsing json which contains gerrit review info, write data to excel.
Changing columns width also delete empty cheets.
Author: s1mple-child
Date: 17-03-2023
'''

import datetime
import json
import sys
import re
import openpyxl
import pandas as pd


def parse_json_from_file(json_file_full_path, target_excel_full_path):
    json_data = []
    for line in open(json_file_full_path, 'rb'):
        json_data.append(json.loads(line))

    for index in range(0, len(json_data) - 1):
        write_change_data_to_excel(target_excel_full_path, json_data[index], index)
        get_review_record_from_patch_set(target_excel_full_path, json_data[index])
        write_subject_to_excel(target_excel_full_path, json_data[index])


def get_basic_information_from_json(source_data):
    project_name = source_data['project']
    branch_name = source_data['branch']
    change_number = source_data['number']
    change_status = source_data['status']
    change_owner = source_data['owner']['username']
    return project_name, branch_name, change_number, change_status, change_owner


def parse_patch_set_from_json(source_data):
    last_patch_set_file_list = []
    last_patch_set_file_type_list = []
    last_patch_set_file_insertions_list = []
    last_patch_set_file_deletions_list = []
    source_patch_set_data = source_data['patchSets']
    patch_set_index_number = len(source_patch_set_data) - 1
    last_patch_set_data = source_patch_set_data[patch_set_index_number]
    last_patch_set_number = last_patch_set_data['number']
    last_patch_set_uploader_name = last_patch_set_data['uploader']['username']
    last_patch_set_created_time = convert_string_to_datetime(last_patch_set_data['createdOn'])

    if 'files' in source_patch_set_data[0]:
        patch_set_files = last_patch_set_data['files']
        for i in range(0, len(patch_set_files)):
            if '/COMMIT_MSG' not in patch_set_files[i]['file']:
                last_patch_set_file_list.append(patch_set_files[i]['file'])
                last_patch_set_file_type_list.append(patch_set_files[i]['type'])
                last_patch_set_file_insertions_list.append(patch_set_files[i]['insertions'])
                last_patch_set_file_deletions_list.append(patch_set_files[i]['deletions'])
    else:
        last_patch_set_file_list.append('No file changed.')
        last_patch_set_file_type_list.append('UNKNOWN')
        last_patch_set_file_insertions_list.append(0)
        last_patch_set_file_deletions_list.append(0)

    project_name, branch_name, change_number, change_status, change_owner = get_basic_information_from_json(source_data)

    return project_name, branch_name, change_number, change_status, change_owner, last_patch_set_number, \
        last_patch_set_uploader_name, last_patch_set_created_time, last_patch_set_file_list, \
        last_patch_set_file_type_list, last_patch_set_file_insertions_list, last_patch_set_file_deletions_list


def get_review_record_from_patch_set(target_excel_full_path, source_data):
    patch_set_reviewed_file_list = []
    patch_set_reviewed_file_line_list = []
    patch_set_reviewer_list = []
    patch_set_review_message_list = []
    patch_set_review_message_severity_level_list = []
    patch_set_review_message_problem_type_list = []
    source_patch_set_data = source_data['patchSets']

    for i in range(0, len(source_patch_set_data)):
        if 'comments' in source_patch_set_data[i]:
            project_name, branch_name, change_number, change_status, change_owner = \
                get_basic_information_from_json(source_data)
            patch_set_number = source_patch_set_data[i]['number']
            patch_set_uploader = source_patch_set_data[i]['uploader']['username']
            last_patch_set_created_time = convert_string_to_datetime(source_patch_set_data[i]['createdOn'])
            patch_set_review_data = source_patch_set_data[i]['comments']
            for j in range(0, len(patch_set_review_data)):
                if patch_set_uploader != patch_set_review_data[j]['reviewer']['username']:
                    patch_set_reviewed_file_list.append(patch_set_review_data[j]['file'])
                    patch_set_reviewed_file_line_list.append(patch_set_review_data[j]['line'])
                    patch_set_reviewer_list.append(patch_set_review_data[j]['reviewer']['username'])
                    patch_set_review_message_list.append(patch_set_review_data[j]['message'])
                    patch_set_review_message_severity_level, patch_set_review_message_problem_type = \
                        parse_inline_message(patch_set_review_data[j]['message'])
                    patch_set_review_message_severity_level_list.append(patch_set_review_message_severity_level)
                    patch_set_review_message_problem_type_list.append(patch_set_review_message_problem_type)
            write_review_data_to_excel(target_excel_full_path, project_name, branch_name, change_status, change_number,
                                       patch_set_number, patch_set_uploader, last_patch_set_created_time,
                                       patch_set_reviewed_file_list, patch_set_reviewed_file_line_list,
                                       patch_set_reviewer_list, patch_set_review_message_list,
                                       patch_set_review_message_severity_level_list,
                                       patch_set_review_message_problem_type_list)


def write_change_data_to_excel(target_excel_full_path, source_data, row_index):
    project_name, branch_name, change_number, change_status, change_owner, patch_set_number, patch_set_uploader, \
        patch_set_created_time, patch_set_file_list, patch_set_file_type_list, \
        patch_set_file_insertions_list, patch_set_file_deletions_list = parse_patch_set_from_json(source_data)

    change_number_with_patch_set_number = str(int(change_number)) + '/' + str(patch_set_number)

    for i in range(0, len(patch_set_file_list)):
        df = pd.DataFrame({'项目名': [project_name],
                           '分支名': [branch_name],
                           'change number/patch set number': [change_number_with_patch_set_number],
                           '状态': [change_status],
                           'Owner': [change_owner],
                           '上传者': [patch_set_uploader],
                           '创建时间': [patch_set_created_time],
                           '文件': [patch_set_file_list[i]],
                           '文件类型': [patch_set_file_type_list[i]],
                           '文件插入行数': [patch_set_file_insertions_list[i]],
                           '文件删除行数': [patch_set_file_deletions_list[i]]}, index=None)
        wb = openpyxl.load_workbook(target_excel_full_path, read_only=True)
        if 'Patch Set' in wb.sheetnames:
            source_df = pd.DataFrame(pd.read_excel(target_excel_full_path, sheet_name='Patch Set'))
            dest_df = pd.concat([source_df, df])
        else:
            dest_df = df

        with pd.ExcelWriter(target_excel_full_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            dest_df.to_excel(writer, index=False, sheet_name='Patch Set')
    print('Add dataframe to excel done. Index number: ' + str(row_index))


def write_review_data_to_excel(target_excel_full_path, project_name, branch_name, change_status, change_number,
                               patch_set_number, patch_set_uploader, created_time, reviewed_file_list,
                               reviewed_file_line_list, reviewed_author_list, reviewed_message_list,
                               inline_message_severity_level_list, inline_message_problem_type_list):
    change_number_with_patch_set_number = str(int(change_number)) + '/' + str(patch_set_number)
    for i in range(0, len(reviewed_file_list)):
        df = pd.DataFrame({'项目名': [project_name],
                           '分支名': [branch_name],
                           'change number/patch set number': [change_number_with_patch_set_number],
                           '状态': [change_status],
                           '上传者': [patch_set_uploader],
                           '创建时间': [created_time],
                           '文件': [reviewed_file_list[i]],
                           '文件行数': [reviewed_file_line_list[i]],
                           '检视人': [reviewed_author_list[i]],
                           '严重级别': [inline_message_severity_level_list[i]],
                           '问题类型': [inline_message_problem_type_list[i]],
                           '检视信息': [reviewed_message_list[i]]}, index=None)
        wb = openpyxl.load_workbook(target_excel_full_path, read_only=True)
        if 'Review Info' in wb.sheetnames:
            source_df = pd.DataFrame(pd.read_excel(target_excel_full_path, sheet_name='Review Info'))
            dest_df = pd.concat([source_df, df])
        else:
            dest_df = df

        with pd.ExcelWriter(target_excel_full_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            dest_df.to_excel(writer, index=False, sheet_name='Review Info')

    print('Add dataframe to sheet Review Info done.')


def write_subject_to_excel(target_excel_full_path, source_data):
    project_name = source_data['project']
    branch_name = source_data['branch']
    change_number = source_data['number']
    change_subject = source_data['subject']

    df = pd.DataFrame({'项目名': [project_name],
                       '分支名': [branch_name],
                       'change number': [change_number],
                       'subject': [change_subject]}, index=None)
    wb = openpyxl.load_workbook(target_excel_full_path, read_only=True)
    if 'Change Subject' in wb.sheetnames:
        source_df = pd.DataFrame(pd.read_excel(target_excel_full_path, sheet_name='Change Subject'))
        dest_df = pd.concat([source_df, df])
    else:
        dest_df = df

    with pd.ExcelWriter(target_excel_full_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        dest_df.to_excel(writer, index=False, sheet_name='Change Subject')

    print('Add subject to sheet Change Subject done.')

def parse_inline_message(inline_message):
    severity_level = ''
    problem_type = ''
    re_pattern = '\\[([a-zA-Z]+)\\]\\[([a-zA-Z]+)\\]'
    result = re.search(re_pattern, inline_message)

    if result is not None:
        severity_level = result.group(1)
        problem_type = result.group(2)
    return severity_level, problem_type

def convert_string_to_datetime(time_string):
    datetime_string = datetime.datetime.fromtimestamp(time_string)
    return str(datetime_string)

parse_json_from_file(sys.argv[1], sys.argv[2])