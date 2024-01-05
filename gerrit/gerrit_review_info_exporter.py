'''
Gerrit Review Data Exporter: Review Info Only

Author: s1mple-child

Step 1: export gerrit data use following command 
'ssh -p 29418 user@gerrit.instance.ip.addr gerrit query --format json --patch-sets --comments --comments \
 --submit-records --files --all-reviewers --no-limit -- -before data-time-u-set > gerrit.json'

Step 2: parse json source file

Step 3: write data to excel or other types of files you want
'''
import json
import datetime
from datetime import datetime, timedelta
import re
import sys
import pandas as pd
import openpyxl
import requests
from openpyxl import load_workbook


# parse json source file and append parsed data to a list
def parse_json_from_file(json_file_full_path):
    """
    :param json_file_full_path: full path of gerrit json file
    :return: json_data: list of data that read from json file
    """
    json_data = []
    # get every line of json file
    for line in open(json_file_full_path, 'rb'):
        json_data.append(json.loads(line))

    return json_data


# get some info from change list like project name, branch name, etc
def get_data_from_change(change_list):
    """
    :param change_list: list that contains change data
    :return: project_name: name of project
    :return: branch_name: name of branch
    :return: change_status: status of change
    :return: change_number: number of change
    """
    project_name = change_list['project']
    branch_name = change_list['branch']
    change_status = change_list['status']
    change_number = change_list['number']

    return project_name, branch_name, change_status, change_number


# get review data use gerrit REST API(https://gerrit-review.googlesource.com/Documentation/rest-api-changes.html#get-comment)
# parse response json and create some lists for writing target excel as data source of pandas dataframe
def find_review_data_per_change(gerrit_instance_ip_addr, change_number):
    """
    :param gerrit_instance_ip_addr: ip address of gerrit
    :param change number: number of change 
    """
    # predefined some lists and append target data to them
    change_reviewed_files_list = []
    change_reviewers_list = []
    change_num_w_patch_set_num_list = []
    change_patch_set_uploaders_list = []
    change_reviewed_line_num_list = []
    change_review_time_list = []
    change_comment_messages_list = []
    inline_message_severity_level_list = []
    inline_message_problem_type_list = []

    # we have to create valid session to request gerrit server
    # so we need a valid user(user with admin permission is recommanded) to achieve this step
    # create HTTP credential of this user in gerrit(Settings - HTTP Credentials)
    gerrit_user_name = 'admin'
    gerrit_http_cred = 'http_passwd'

    # create session of user
    session = requests.Session()
    # auth with defined gerrit user namd and http cred
    session.auth = (gerrit_user_name, gerrit_http_cred)
    # request with specific change number via above API address
    # at the beginning of response contains some extra brackets, delete them with replace
    comment_response = session.get('http://' + gerrit_instance_ip_addr + '/a/changes/' + str(change_number) 
                                   + '/comments').text.replace(')]}\'', '')
    # convert response text to json
    comment_response_json = json.loads(comment_response)
    # parse json response
    # iterate key in json
    for key in comment_response_json.keys():
        for i in range(len(comment_response_json[key])):
            # patch set number
            patch_set_number = comment_response_json[key][i]['patch_set']
            # find patch set uploader use patch_set_number
            patch_set_uploader = find_patch_set_uploader(gerrit_instance_ip_addr, change_number, patch_set_number)
            # get review data that reviewer name who commented is different with patch set uploader name
            if comment_response_json[key][i]['author']['name'] != patch_set_uploader:
                # create some lists as source data export to excel
                # reviewed files list
                change_reviewed_files_list.append(key)
                # reviewers list
                reviewer = comment_response_json[key][i]['author']['username']
                change_reviewers_list.append(reviewer)
                # change number and patch set number list
                change_num_w_patch_set_num = str(change_number) + '/' + str(patch_set_number)
                change_num_w_patch_set_num_list.append(change_num_w_patch_set_num)
                # patch set uploader list
                change_patch_set_uploaders_list.append(patch_set_uploader)
                # reviewed lines number list
                # check 'line' in response json set empty if not exist
                if 'line' in comment_response_json[key][i]['line']:
                    reviewed_line_num = comment_response_json[key][i]['line']
                else:
                    reviewed_line_num = ''
                change_reviewed_line_num_list.append(reviewed_line_num)
                # find review time and append to review time list
                # NOTE: result time format is yyyy-MM-dd:S000f (with 9 microseconds) and timezone is UTC
                # remove microseconds before set value to raw_review_time, then change timezone to local timezone use datetime.timedelta
                raw_review_time = datetime.strptime(comment_response_json[key][i]['updated'], '%Y-%m-%d %H:%M:%S.000%f').replace(microsecond=0)
                # change raw_review_time timezone to local timezone
                local_review_time = raw_review_time + timedelta(hours=8)
                change_review_time_list.append(local_review_time)
                # comment messages list
                comment_message = comment_response_json[key][i]['message']
                change_comment_messages_list.append(comment_message)
                # in my use case, inline messages have some specific pattern
                # inline messages pattern: [RE,G]: blablabla
                # RE means code error type about misunstanding the requirement 
                # G means the severity level is normal
                # so export data also parse these inline messages and export above these two elements to separate columns
                inline_message_severity_level, inline_message_problem_type = parse_inline_messages(comment_message)
                inline_message_severity_level_list.append(inline_message_severity_level)
                inline_message_problem_type_list.append(inline_message_problem_type)
    
    return (change_reviewed_files_list, change_reviewers_list, change_num_w_patch_set_num_list, change_patch_set_uploaders_list, 
            change_reviewed_line_num_list, change_review_time_list, change_comment_messages_list, inline_message_severity_level_list, 
            inline_message_problem_type_list)


# write review data with several lists created in find_review_data_per_change to target file
# we also use date as parameters to filter data between two dates and write these data to target
def write_result_lists_to_target(start_date, end_date, reviewed_date_list, 
                                patch_set_num_list, reviewed_messages_list, reviewed_file_list, 
                                line_num_reviewed_file_list, liline_message_problem_type_list, 
                                inline_message_severity_level_list, change_status, patch_set_uploader_list, 
                                reviewers_list, project_name, branch_name, target_file_full_path):
    # the format of start_date and end_date is YYYY-mm-dd
    # convert two string date values to datetime format
    formatted_start_date = datetime.strptime(start_date, '%Y-%m-%d')
    formatted_end_date = datetime.strptime(end_date, '%Y-%m-%d')

    # because all given lists have same length, select one of them and itearate all given lists
    for i, key in enumerate(reviewed_file_list):
        if formatted_start_date <= key <= formatted_end_date:
            print('Writing review data to target excel which sheet name is Review Info. The ref is ' + patch_set_num_list[i])
            # create pandas dataframe
            df = pd.DataFrame({'review message': [reviewed_messages_list[i]],
                                'review date': [reviewed_date_list[i]],
                                'reviewed file': [reviewed_file_list[i]],
                                'line number of reviewed file': [line_num_reviewed_file_list[i]],
                                'inline message problem type': [liline_message_problem_type_list[i]],
                                'inline message severity level': [inline_message_severity_level_list[i]],
                                'change status': [change_status],
                                'patch set uploader': [patch_set_uploader_list[i]],
                                'reviewer': [reviewers_list[i]],
                                'project name': [project_name],
                                'branch name': [branch_name],
                                'change number/patch set number': [patch_set_num_list[i]]
                                })
            # load workbook from given target file full path
            wb = openpyxl.load_workbook(target_file_full_path, read_only=True)
            if 'Review Info' in wb.sheetnames:
                # use concat to concreate existed data and new data
                source_df = pd.DataFrame(pd.read_excel(target_file_full_path, sheet_name='Review Info'))
                target_df = pd.concat([source_df, df])
            else:
                target_df = df
            # write new dataframe object to target file
            with pd.ExcelWriter(target_file_full_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                target_df.to_excel(writer, index=False, sheet_name='Review Info')


# the entrance of file
def parse_n_write_data_to_target(json_file_full_path, target_file_full_path, gerrit_instance_ip_addr, start_date, end_date):
    """
    :param: json_file_full_path: full path of gerrit source json file
    :param: target_file_full_path: full path of target file
    :param: gerrit_instance_ip_addr: gerrit ip address
    :param: start_date: start date
    :param: end_date: end date
    """
    source_gerrit_json_data = parse_json_from_file(json_file_full_path)
    for index in range(len(source_gerrit_json_data) - 1):
        # get data of change
        change_data = source_gerrit_json_data[index]
        project_name = get_data_from_change(change_data)[0]
        branch_name = get_data_from_change(change_data)[1]
        change_status = get_data_from_change(change_data)[2]
        change_number = get_data_from_change(change_data)[3]

        # get all lists of change
        patch_set_data = source_gerrit_json_data[index]['patchSets']
        for count, key in enumerate(patch_set_data):
            if 'comments' in key:
                change_reviewed_files_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[0]
                change_reviewers_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[1]
                change_patch_set_num_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[2]
                change_patch_set_uploaders_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[3]
                change_reviewed_line_num_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[4]
                change_review_time_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[5]
                change_comment_messages_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[6]
                inline_message_severity_level_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[7]
                inline_message_problem_type_list = find_review_data_per_change(gerrit_instance_ip_addr, change_number)[8]
                # write results to target file
                write_result_lists_to_target(start_date, end_date, change_review_time_list, change_patch_set_num_list, 
                                             change_comment_messages_list, change_reviewed_files_list, change_reviewed_line_num_list, 
                                             inline_message_problem_type_list, inline_message_severity_level_list, change_status, 
                                             change_patch_set_uploaders_list, change_reviewers_list, project_name, branch_name, 
                                             target_file_full_path)
                break


# find every patch set uploader of change
# because other users can rebase or cherry pick which change is not belong to 'himself/herself'
# this will make difference between patch set uploader and owner of a change
# we will request via another API(https://gerrit-review.googlesource.com/Documentation/rest-api-changes.html#get-change-detail)
def find_patch_set_uploader(gerrit_instance_ip_addr, change_number, patch_set_number):
    """
    :param gerrit_instance_ip_addr: ip address of gerrit
    :param change_number: number of change
    :param patch_set_number: number of patchset
    :return: patch_set_uploader: uploader of patchset
    """
    gerrit_user_name = 'admin'
    gerrit_http_cred = 'http_passwd'
    # request via API
    session = requests.Session()
    session.auth = (gerrit_user_name, gerrit_http_cred)
    change_details_response = session.get('http://' + gerrit_instance_ip_addr + '/a/changes/' 
                                          + str(change_number) + '/detail').text.replace(')]}\'', '')
    change_details_response_json = json.loads(change_details_response)

    # the patch set uploader located in messages of response json
    for j in range(len(change_details_response_json['messages'])):
        # compare patch set number in response json with patch_set_number
        if change_details_response_json['messages'][j]['_revision_number'] == patch_set_number and 'tag' \
            in change_details_response_json['messages'][j]:
            # we regard a tag that conatins 'gerrit:newPatchSet' as the first record of every patch set
            if 'gerrit:newPatchSet' in change_details_response_json['messages'][j]['tag']:
                # get patch set uploader
                patch_set_uploader = change_details_response_json['messages'][j]['author']['name']

    return patch_set_uploader


# parse inline messages: find specific codes in every inline message and convert these codes to its details
def parse_inline_messages(inline_message):
    """
    :param inline_message: inline message
    :return: actually_mean_severity_level: the complete meaning of severity level
    :return: actually_mean_problem_type: the complete meaning of problem type
    """
    # set inline message pattern and check if inline message is corresponding its pattern
    re_pattern = '\\[([a-zA-Z]+),([a-zA-Z]+)\\]'
    check_pattern_result = re.search(re_pattern, inline_message)

    if check_pattern_result is not None:
        problem_type = check_pattern_result.group(1)
        severity_level = check_pattern_result.group(2)
        if problem_type == 'RE':
            actually_mean_problem_type = 'error of understanding requirement'
        elif problem_type == 'LE':
            actually_mean_problem_type = 'error of running result'
        elif problem_type == 'CO':
            actually_mean_problem_type = 'code can promote'
        elif problem_type == 'CS':
            actually_mean_problem_type = 'syntax error'
        elif problem_type == 'CE':
            actually_mean_problem_type = 'error of code instruction'

        if severity_level == 'F':
            actually_mean_severity_level = 'fatal'
        elif severity_level == 'S':
            actually_mean_severity_level = 'critical'
        elif severity_level == 'G':
            actually_mean_severity_level = 'major'
        elif severity_level == 'T':
            actually_mean_severity_level = 'minor'

    return actually_mean_severity_level, actually_mean_problem_type


# insert index number in first column and set border in target file
def write_index_n_set_border(target_file_full_path):
    """
    :param target_file_full_path: full path of target file
    """
    full_df = pd.DataFrame(pd.read_excel(target_file_full_path, sheet_name='Review Info'))
    full_df.index += 1
    with pd.ExcelWriter(target_file_full_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        full_df.style.set_properties(**{'text-align': 'left', 'border-color': 'black', 'border-width': '1px', 
        'border-style': 'solid'}).to_excel(writer, index_label='Index', sheet_name='Review Info')


# set width of all columns
def set_columns_width(target_file_full_path):
    """
    :param target_file_full_path: full path of target file
    """
    wb = load_workbook(filename=target_file_full_path)
    for sheet_name in wb.sheetnames:
        for col in wb[sheet_name].columns:
            # find title of every column
            column = col[0].column_letter
            if sheet_name == 'Review Info':
                # set column to a specific width number
                if column == 'A':
                    wb[sheet_name].column_dimensions[column].width = 5
                elif column in('B', 'D'):
                    wb[sheet_name].column_dimensions[column].width = 55
                elif column in ('E', 'G'):
                    wb[sheet_name].column_dimensions[column].width = 10
                elif column in ('F', 'H', 'I', 'J'):
                    wb[sheet_name].column_dimensions[column].width = 15
                else:
                    wb[sheet_name].column_dimensions[column].width = 23
    """
    Method2: set column width depend on the longest width of cell
    for col in wb[sheet_name].columns:
        max_width = 20
        adjusted_width = 0
        column = col[0].column_letter
        for cell in col:
            print(str(cell.value) + ': ' + str(len(str(cell.value))))
            # compare width with max_width
            # set adjusted_width if width of cell is bigger than max_width
            if len(str(cell.value)) > max_length:
                max_width = 28
                adjusted_width = max_width + 2
            # set column width to adjusted_width's value
            wb[sheet_name].column_dimensions[column].width = adjusted_width
    """
    wb.save(target_file_full_path)


# delete empty sheet
def delete_empty_sheet(target_file_full_path):
    """
    :param target_file_full_path: full path of target file
    """
    wb = load_workbook(filename=target_file_full_path)
    for sheet_name in wb.sheetnames:
        if wb[sheet_name].max_column == 1:
            del wb[sheet_name]
        else:
            print('Sheet ' + sheet_name + ' not empty, skip.')

    wb.save(target_file_full_path)


'''
:parameter
1. source gerrit json full path
2. target file full path
3. gerrit instance ip address
4. start date(format: YYYY-mm-dd)
5. end date(format: YYYY-mm-dd)
'''
parse_n_write_data_to_target(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5])
write_index_n_set_border(sys.argv[2])
set_columns_width(sys.argv[2])
delete_empty_sheet(sys.argv[2])
